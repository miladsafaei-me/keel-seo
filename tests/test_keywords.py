"""Tests for keel_seo.keywords -- the autocomplete keyword-universe crawler.

Plain unittest, no Django and no network: the collector is driven by a stub that
replays canned responses, so the crawl, the scoring and the clustering are all
tested against fixed inputs.

Run: DJANGO_SETTINGS_MODULE=tests.settings python -m django test tests.test_keywords
     python -m unittest tests.test_keywords
"""
import os
import tempfile
import unittest
from pathlib import Path

from keel_seo.keywords import cluster as clustering
from keel_seo.keywords import harvest, report, sync
from keel_seo.keywords import language
from keel_seo.keywords import markets as target_markets
from keel_seo.keywords.crawl import (PROBE_NOVELTY_FLOOR, PROBE_NOVELTY_SHARE,
                                     PROBE_QUERIES, Phrase, Universe,
                                     contains_seed, crawl, discover_variants,
                                     merge_markets, probe_market, score,
                                     seed_spelling, seed_tokens, squash,
                                     worth_crawling)
from keel_seo.keywords.grammar import (BRANCH, DRILL, SEED, expansions,
                                       star_variants)
from keel_seo.keywords.proxying import (DIRECT_REFUSAL, DirectEgressRefused,
                                        accept_suggestions,
                                        require_pooled_egress)
from keel_seo.keywords.suggest import Response, SuggestCache, SuggestClient, Suggestion


class StubClient(SuggestClient):
    """A SuggestClient that answers from a dict instead of the network."""

    def __init__(self, answers, capacity=15):
        super().__init__(cache=SuggestCache(None))
        self.answers = answers
        self._capacity = capacity
        self.asked = []

    @property
    def capacity(self):
        return self._capacity

    def fetch(self, query):
        self.asked.append(query)
        phrases = self.answers.get(query, [])
        return Response(
            query=query,
            suggestions=tuple(
                Suggestion(p, i, 600 - i) for i, p in enumerate(phrases, 1)
            ),
            capacity=self._capacity,
        )


class GrammarTests(unittest.TestCase):
    def test_seed_tier_surrounds_the_term_on_both_sides(self):
        queries = set(expansions("quotex", SEED))
        self.assertIn("quotex a", queries, "spaced suffix sweep missing")
        self.assertIn("quotex 9", queries, "digit suffix missing")
        self.assertIn("quotexa", queries, "tight (no-space) suffix missing")
        self.assertIn("a quotex", queries, "spaced prefix sweep missing")
        self.assertIn("is quotex", queries, "prefix word missing")
        self.assertIn("quotex vs", queries, "suffix word missing")
        self.assertIn("quotex", queries, "the bare seed must be asked too")

    def test_a_star_walks_every_gap_between_words(self):
        """The family that reaches phrases where the term is not the leading text."""
        self.assertEqual(star_variants("quotex signal bot"),
                         ["quotex * signal bot", "quotex signal * bot"])

    def test_a_single_word_term_has_no_gap_to_walk(self):
        self.assertEqual(star_variants("quotex"), [])

    def test_the_seed_tier_asks_everything_twice_with_a_leading_space(self):
        queries = expansions("quotex", SEED)
        self.assertIn(" quotex", queries, "a leading space is its own family")
        self.assertIn(" quotex a", queries)
        self.assertEqual(sum(1 for q in queries if q.startswith(" ")), len(queries) // 2)

    def test_a_trailing_space_is_never_asked(self):
        """Google trims it, so the query is byte-identical to the bare one."""
        for tier in (SEED, BRANCH, DRILL):
            self.assertFalse(any(q.endswith(" ") for q in expansions("quotex bot", tier)),
                             f"{tier} emitted a trailing-space query, which is a "
                             "duplicate request for identical data")

    def test_branch_phrases_get_star_variants(self):
        queries = expansions("quotex signal bot", BRANCH)
        self.assertIn("quotex * signal bot", queries)

    def test_tight_sweep_can_be_switched_off(self):
        self.assertNotIn("quotexa", expansions("quotex", SEED, tight=False))

    def test_branch_tier_is_cheaper_than_seed_tier(self):
        self.assertLess(
            len(expansions("quotex signal bot", BRANCH)),
            len(expansions("quotex signal bot", SEED)),
        )

    def test_drill_tier_is_the_alphabet_only(self):
        self.assertEqual(len(expansions("quotex app", DRILL)), 26)

    def test_no_query_is_asked_twice(self):
        queries = expansions("quotex", SEED)
        self.assertEqual(len(queries), len(set(queries)))

    def test_empty_term_yields_nothing(self):
        self.assertEqual(expansions("   ", SEED), [])


class ContainmentTests(unittest.TestCase):
    def test_compound_without_space_still_contains_the_seed(self):
        self.assertTrue(contains_seed("quotexapk", seed_tokens("quotex")))

    def test_multi_word_seed_matches_in_any_order(self):
        tokens = seed_tokens("pip value calculator")
        self.assertTrue(contains_seed("calculator for pip value", tokens))
        self.assertFalse(contains_seed("pip calculator", tokens))

    def test_off_topic_neighbour_is_rejected(self):
        self.assertFalse(contains_seed("how to quote on reddit", seed_tokens("quotex")))


class SaturationTests(unittest.TestCase):
    def test_full_response_is_saturated_and_short_one_is_not(self):
        full = Response("q", tuple(Suggestion(f"p{i}", i, 1) for i in range(15)), 15)
        short = Response("q", tuple(Suggestion(f"p{i}", i, 1) for i in range(4)), 15)
        self.assertTrue(full.saturated)
        self.assertFalse(short.saturated)

    def test_parser_reads_relevance_and_survives_its_absence(self):
        with_scores = SuggestClient._parse(
            ["q", ["a", "b"], ["", ""], [], {"google:suggestrelevance": [900, 800]}]
        )
        self.assertEqual(with_scores, [["a", 900], ["b", 800]])
        self.assertEqual(SuggestClient._parse(["q", ["a"]]), [["a", 0]])
        self.assertEqual(SuggestClient._parse(None), [])


class CrawlTests(unittest.TestCase):
    def test_only_seed_bearing_phrases_enter_the_universe(self):
        client = StubClient({"quotex": ["quotex login", "how to quote on reddit"]})
        universe = crawl("quotex", client, levels=0, saturate=0, budget=500)
        self.assertIn("quotex login", universe.phrases)
        self.assertNotIn("how to quote on reddit", universe.phrases)
        self.assertEqual(universe.off_seed["how to quote on reddit"], 1)

    def test_reach_counts_distinct_parent_queries(self):
        client = StubClient({
            "quotex": ["quotex app"],
            "quotex a": ["quotex app"],
            "quotex b": ["quotex app"],
        })
        universe = crawl("quotex", client, levels=0, saturate=0, budget=500)
        self.assertEqual(universe.phrases["quotex app"].reach, 3)

    def test_budget_is_a_hard_cap(self):
        client = StubClient({})
        universe = crawl("quotex", client, levels=2, saturate=1, budget=25)
        self.assertLessEqual(universe.queries_asked, 25)

    def test_drilling_happens_only_under_a_truncated_response(self):
        client = StubClient({"quotex": ["quotex a" + c for c in "abcd"]}, capacity=15)
        crawl("quotex", client, levels=0, saturate=1, budget=500)
        self.assertNotIn("quotex a", [q for q in client.asked if q.endswith(" a")][1:],
                         "a short response must not trigger a drill round")

    def test_a_deadline_stops_the_crawl_but_keeps_what_it_found(self):
        """An external kill loses everything; the crawl's own deadline loses nothing."""
        client = StubClient({"quotex": ["quotex login", "quotex app"]})
        universe = crawl("quotex", client, levels=2, saturate=0, budget=5000,
                         max_seconds=0.001)
        self.assertTrue(universe.timed_out)
        self.assertFalse(universe.exhausted)

    def test_no_deadline_by_default(self):
        client = StubClient({"quotex": ["quotex login"]})
        universe = crawl("quotex", client, levels=0, saturate=0, budget=500)
        self.assertFalse(universe.timed_out)

    def test_a_closed_universe_reports_itself_exhausted(self):
        client = StubClient({})
        universe = crawl("quotex", client, levels=1, saturate=0, budget=500)
        self.assertTrue(universe.exhausted)
        self.assertEqual(universe.unexpanded, 0)


class RateLimitTests(unittest.TestCase):
    """The endpoint answers 403 when pushed, and says nothing else about it."""

    def test_throttle_spaces_requests_across_threads(self):
        from concurrent.futures import ThreadPoolExecutor
        import time

        from keel_seo.keywords.suggest import Throttle

        throttle = Throttle(rate=50.0)
        started = time.monotonic()
        with ThreadPoolExecutor(max_workers=5) as pool:
            list(pool.map(lambda _: throttle.wait(), range(10)))
        self.assertGreaterEqual(time.monotonic() - started, 0.9 * 9 / 50.0)

    def test_zero_rate_disables_the_throttle(self):
        from keel_seo.keywords.suggest import Throttle

        self.assertEqual(Throttle(rate=0).interval, 0.0)

    def test_a_block_code_trips_the_breaker_and_a_failure_does_not(self):
        from keel_seo.keywords.suggest import BLOCK_CODES, BLOCK_LIMIT

        self.assertIn(403, BLOCK_CODES, "Google answers 403, not 429, for this endpoint")
        client = SuggestClient(cache=SuggestCache(None), rate=0)
        for _ in range(BLOCK_LIMIT - 1):
            client._note_block()
        self.assertFalse(client.blocked)
        client._note_block()
        self.assertTrue(client.blocked)

    def test_one_success_clears_an_isolated_run_of_blocks(self):
        client = SuggestClient(cache=SuggestCache(None), rate=0)
        client._note_block()
        client._note_block()
        client._note_success()
        client._note_block()
        self.assertFalse(client.blocked)

    def test_a_cache_from_another_country_is_not_reused(self):
        """Autocomplete answers by IP, so geography is part of a response's identity."""
        turkey = SuggestCache(None, egress="TR")
        germany = SuggestCache(None, egress="DE")
        self.assertNotEqual(turkey.key("chrome", "en", "", "quotex"),
                            germany.key("chrome", "en", "", "quotex"))
        turkey.put(turkey.key("chrome", "en", "", "quotex"), [["quotex login", 900]])
        self.assertIsNone(germany.get(germany.key("chrome", "en", "", "quotex")))

    def test_language_and_vertical_are_also_part_of_the_key(self):
        cache = SuggestCache(None, egress="TR")
        base = cache.key("chrome", "en", "", "quotex")
        self.assertNotEqual(base, cache.key("chrome", "pt-BR", "", "quotex"))
        self.assertNotEqual(base, cache.key("chrome", "en", "yt", "quotex"))
        self.assertNotEqual(base, cache.key("firefox", "en", "", "quotex"))

    def test_a_blocked_client_stops_making_requests(self):
        client = SuggestClient(cache=SuggestCache(None), rate=0)
        client.blocked = True
        response = client.fetch("quotex")
        self.assertEqual(response.error, "blocked")
        self.assertEqual(client.calls, 0)

    def test_a_blocked_crawl_keeps_what_it_already_collected(self):
        class Blocking(StubClient):
            def fetch(self, query):
                response = super().fetch(query)
                if len(self.asked) > 3:
                    self.blocked = True
                return response

        client = Blocking({"quotex": ["quotex login", "quotex app"]})
        universe = crawl("quotex", client, levels=2, saturate=0, budget=5000)
        self.assertTrue(universe.blocked)
        self.assertIn("quotex login", universe.phrases)
        self.assertLess(universe.queries_asked, 5000,
                        "a blocked crawl must stop, not spend the whole budget")


class CommandLineTests(unittest.TestCase):
    """The CLI is a code path too, and it broke while every other test passed.

    Removing cluster.DEFAULT_THRESHOLD left __main__ referencing it, which no test
    caught because they all call build() directly. The harvest died on the server
    with AttributeError after a clean local run.
    """

    def test_the_parser_builds_and_its_defaults_resolve(self):
        from keel_seo.keywords.__main__ import build_parser

        args = build_parser().parse_args(["quotex"])
        self.assertEqual(args.topics, clustering.DEFAULT_TOPICS)
        self.assertEqual(args.seed, "quotex")

    def test_every_clustering_option_the_cli_offers_is_accepted_by_build(self):
        import inspect

        from keel_seo.keywords.__main__ import build_parser

        args = build_parser().parse_args(["quotex", "--topics", "12"])
        accepted = inspect.signature(clustering.build).parameters
        self.assertIn("topics", accepted,
                      "the CLI passes --topics; build() must take it")
        universe = Universe(seed="quotex")
        from keel_seo.keywords.crawl import Phrase

        for text in ("quotex demo one", "quotex demo two", "quotex demo three"):
            universe.phrases[text] = Phrase(text, 1, 600, 0)
        score(universe)
        clustering.build(universe, topics=args.topics)   # must not raise

    def test_names_the_cli_reads_off_the_clustering_module_exist(self):
        for name in ("DEFAULT_TOPICS", "TAIL_LABEL", "build"):
            self.assertTrue(hasattr(clustering, name), f"cluster.{name} is gone")


class WorkbookTests(unittest.TestCase):
    """The .xlsx is optional, and its absence must cost only the workbook."""

    def universe(self):
        from keel_seo.keywords.crawl import Phrase

        u = Universe(seed="quotex")
        for i, text in enumerate(["quotex login", "quotex login page", "quotex apk"]):
            phrase = Phrase(text, 1 + i, 600 - i, 0)
            phrase.parents = {f"q{i}"}
            u.phrases[text] = phrase
        u.off_seed = {"quotes meaning in urdu": 4}
        score(u)
        return u

    def _workbook(self):
        import tempfile

        from keel_seo.keywords import report

        universe = self.universe()
        clusters = clustering.build(universe)
        meta = report.metadata(universe, clusters,
                               {"ip": "", "country": "mixed", "org": "pool"},
                               SuggestClient(cache=SuggestCache(None), rate=0))
        tmp = tempfile.mkdtemp()
        path = os.path.join(tmp, "wb.xlsx")
        if not report.write_xlsx(path, universe, clusters, meta):
            self.skipTest("openpyxl not installed")
        import openpyxl

        return openpyxl.load_workbook(path), clusters

    def test_run_is_the_first_sheet(self):
        book, _ = self._workbook()
        self.assertEqual(book.sheetnames[0], "Run",
                         "the reader needs provenance before any number")

    def test_run_carries_a_summary_not_the_whole_metadata_dump(self):
        book, _ = self._workbook()
        labels = [r[0] for r in book["Run"].iter_rows(values_only=True)]
        self.assertNotIn("volume_note", labels)
        self.assertNotIn("proxy_pool", labels)
        self.assertIn("Keywords found", labels)
        self.assertIn("How it ended", labels)

    def test_the_summary_count_matches_the_keywords_sheet(self):
        """A Run sheet that disagrees with the sheet beside it is worse than none."""
        book, _ = self._workbook()
        rows = {r[0]: r[1] for r in book["Run"].iter_rows(values_only=True)}
        self.assertEqual(rows["Keywords found"].replace(",", ""),
                         str(book["Keywords"].max_row - 1))

    def test_run_explains_every_scored_column(self):
        book, _ = self._workbook()
        labels = {r[0] for r in book["Run"].iter_rows(values_only=True)}
        for column in ("Priority", "Reach", "Relevance", "Level", "Best rank"):
            self.assertIn(column, labels, f"{column} is unexplained")

    def test_the_volume_caveat_travels_with_priority(self):
        book, _ = self._workbook()
        text = " ".join(str(r[1]) for r in book["Run"].iter_rows(values_only=True)
                        if r[0] == "Priority").lower()
        # The concept, not a form of words: this must survive rewording, and
        # asserting an exact phrase made an editing pass look like a regression.
        self.assertIn("search volume", text,
                      "Priority must say it is not a volume figure")

    def test_every_cluster_row_links_to_its_own_keywords(self):
        book, clusters = self._workbook()
        sheet, keywords = book["Clusters"], book["Keywords"]
        # Find the cluster column by its header, not by position: a new column
        # inserted before it should not make this test fail for the wrong reason.
        from openpyxl.utils import get_column_letter

        headers = {c.value: get_column_letter(c.column) for c in keywords[1]}
        cluster_col = headers["Cluster"]
        for row in range(2, sheet.max_row + 1):
            link = sheet[f"A{row}"].hyperlink
            self.assertIsNotNone(link, f"cluster row {row} is not clickable")
            # location, not target: an in-workbook jump carries no external target.
            self.assertIsNone(link.target, "an internal jump must have no target")
            self.assertTrue(link.location.endswith(tuple("0123456789")))
            # The link must land on a row that really belongs to that cluster.
            target_row = int(link.location.split("A")[-1])
            self.assertEqual(keywords[f"{cluster_col}{target_row}"].value,
                             sheet[f"A{row}"].value)

    def test_cluster_links_are_internal_jumps_not_external_targets(self):
        """The bug LibreOffice exposed and Excel had been hiding.

        Assigning a plain "#Sheet!A1" string makes openpyxl emit an EXTERNAL
        relationship whose target starts with "#". That is malformed for an
        in-workbook jump: LibreOffice ignored it entirely (603 links, 0 internal)
        while Excel resolved it only by leniency. The standard form is a
        `location` attribute and no relationship at all, so the XML is asserted
        here rather than the openpyxl object, which reports both cases the same.
        """
        import re
        import tempfile
        import zipfile

        from keel_seo.keywords import report

        universe = self.universe()
        clusters = clustering.build(universe, min_anchor=3)
        meta = report.metadata(universe, clusters,
                               {"ip": "", "country": "mixed", "org": "pool"},
                               SuggestClient(cache=SuggestCache(None), rate=0))
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "wb.xlsx")
            if not report.write_xlsx(path, universe, clusters, meta):
                self.skipTest("openpyxl not installed")
            archive = zipfile.ZipFile(path)
            tags = []
            for name in archive.namelist():
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                    # The trailing space matters: <hyperlinks> is the container,
                    # not a link, and would fail every assertion below.
                    tags += re.findall(r"<hyperlink [^>]*>",
                                       archive.read(name).decode("utf-8"))
            self.assertTrue(tags, "the Clusters sheet should carry links")
            for tag in tags:
                self.assertIn("location=", tag)
                self.assertNotIn("r:id=", tag,
                                 "an internal jump must not go through a relationship")
            rels = [n for n in archive.namelist()
                    if "worksheets/_rels" in n and "hyperlink" in archive.read(n).decode()]
            self.assertEqual(rels, [], "internal jumps need no relationship file")

    def test_the_other_three_formats_survive_a_missing_openpyxl(self):
        import builtins
        import tempfile

        from keel_seo.keywords import report

        universe = self.universe()
        clusters = clustering.build(universe)
        real_import = builtins.__import__

        def no_openpyxl(name, *args, **kwargs):
            if name.startswith("openpyxl"):
                raise ImportError("simulated: openpyxl not installed")
            return real_import(name, *args, **kwargs)

        # Real metadata, not a stub: write_markdown reads a dozen fields and a
        # half-filled dict would fail this test for the wrong reason.
        meta = report.metadata(universe, clusters,
                               {"ip": "1.2.3.4", "country": "TR", "org": "test"},
                               SuggestClient(cache=SuggestCache(None), rate=0))

        with tempfile.TemporaryDirectory() as tmp:
            builtins.__import__ = no_openpyxl
            try:
                paths = report.write_all(tmp, universe, clusters, meta)
            finally:
                builtins.__import__ = real_import
            self.assertNotIn("xlsx", paths, "no workbook without the library")
            for kind in ("json", "csv", "md"):
                self.assertIn(kind, paths)
                self.assertTrue(os.path.getsize(paths[kind]) > 0,
                                f"{kind} must still be written")


class LanguageTests(unittest.TestCase):
    """Three tiers of certainty, and the false positives that shaped tier three."""

    def test_a_non_latin_script_is_certain(self):
        self.assertEqual(language.non_english_reason("quotex telegram отзывы"),
                         language.SCRIPT)
        self.assertEqual(language.non_english_reason("quotex 日本"), language.SCRIPT)

    def test_accented_letters_are_near_certain(self):
        self.assertEqual(language.non_english_reason("quotex app móvil"),
                         language.DIACRITIC)
        self.assertEqual(language.non_english_reason("quotex binäre option"),
                         language.DIACRITIC)

    def test_a_marker_word_names_itself_so_the_call_can_be_checked(self):
        reason = language.non_english_reason("descargar quotex apk")
        self.assertTrue(reason.startswith(language.VOCABULARY))
        self.assertIn("descargar", reason)

    def test_plain_english_is_left_alone(self):
        for text in ("quotex demo account", "quotex app download apk",
                     "is quotex legit", "quotex withdrawal problem"):
            self.assertEqual(language.non_english_reason(text), "", text)

    def test_the_two_false_positives_that_shaped_the_marker_list(self):
        """A domain part and a crypto term, both flagged by earlier drafts."""
        self.assertEqual(language.non_english_reason("quotex com login"), "",
                         "'com' is a domain, not Portuguese")
        self.assertEqual(language.non_english_reason("quotex dao download"), "",
                         "'dao' here is the crypto sense, not Vietnamese")

    def test_no_marker_is_itself_an_ordinary_english_word(self):
        english = {"app", "com", "download", "demo", "login", "free", "best",
                   "account", "bot", "signal", "trading", "broker", "review",
                   "legal", "safe", "real", "code", "bonus", "dao", "para"}
        self.assertEqual(language.MARKER_WORDS & english, set(),
                         "a marker that is also English costs more than it finds")


class ProxySeamTests(unittest.TestCase):
    """What stays here: the endpoint-specific half of proxy verification.

    Rotation, the durable store and its ageing policy belong to keel-crawler
    (`keel_crawler.proxy.pool`) and are tested there — this package only decides
    what a usable answer from *this* endpoint looks like.
    """

    def test_a_real_autocomplete_answer_is_accepted(self):
        self.assertTrue(accept_suggestions(200, '["quotex",["quotex login"]]'))

    def test_a_captive_portal_is_rejected_despite_its_200(self):
        self.assertFalse(accept_suggestions(200, "<html><body>Login required</body></html>"),
                         "a proxy returning an interstitial would otherwise join the "
                         "pool and then fail every real request")

    def test_a_refusal_or_an_empty_body_is_rejected(self):
        self.assertFalse(accept_suggestions(403, "[]"))
        self.assertFalse(accept_suggestions(200, "   "))


class ScoreTests(unittest.TestCase):
    def build(self, rows):
        universe = Universe(seed="quotex")
        from keel_seo.keywords.crawl import Phrase

        for text, rank, relevance, level, parents in rows:
            phrase = Phrase(text, rank, relevance, level)
            phrase.parents = set(parents)
            universe.phrases[text] = phrase
        score(universe)
        return universe

    def test_better_rank_and_wider_reach_outrank_the_tail(self):
        universe = self.build([
            ("quotex login", 1, 900, 0, ["a", "b", "c", "d"]),
            ("quotex zigzag settings pdf", 14, 550, 2, ["z"]),
        ])
        self.assertGreater(
            universe.phrases["quotex login"].priority,
            universe.phrases["quotex zigzag settings pdf"].priority,
        )

    def test_ranked_output_is_ordered_by_priority(self):
        universe = self.build([
            ("quotex tail", 12, 550, 2, ["z"]),
            ("quotex head", 1, 900, 0, ["a", "b"]),
        ])
        self.assertEqual([p.text for p in universe.ranked()],
                         ["quotex head", "quotex tail"])


class ClusterTests(unittest.TestCase):
    def universe(self, phrases):
        from keel_seo.keywords.crawl import Phrase

        universe = Universe(seed="quotex")
        for index, text in enumerate(phrases):
            phrase = Phrase(text, 1 + index % 5, 600 - index, 0)
            phrase.parents = {f"p{index}"}
            universe.phrases[text] = phrase
        score(universe)
        return universe

    def corpus(self):
        """A corpus with real topics: words that name a group, and words that do not."""
        return [
            "quotex zigzag strategy", "quotex zigzag indicator", "quotex zigzag settings",
            "quotex withdrawal problem", "quotex withdrawal proof", "quotex withdrawal time",
            "quotex demo account", "quotex demo login", "quotex demo reset",
            "quotex martingale strategy", "quotex rsi strategy", "quotex trend strategy",
        ]

    def test_a_keyword_lands_under_the_word_that_says_what_it_is_about(self):
        clusters = clustering.build(self.universe(self.corpus()), min_anchor=3)
        by_phrase = {p.text: c for c in clusters for p in c.members}
        self.assertEqual(by_phrase["quotex zigzag strategy"].index,
                         by_phrase["quotex zigzag indicator"].index)
        self.assertEqual(by_phrase["quotex withdrawal problem"].index,
                         by_phrase["quotex withdrawal proof"].index)
        self.assertNotEqual(by_phrase["quotex zigzag strategy"].index,
                            by_phrase["quotex withdrawal proof"].index)

    def test_the_rarer_word_wins_when_a_phrase_holds_two_topics(self):
        """`zigzag strategy` is about zigzag; `strategy` is spread across the corpus."""
        clusters = clustering.build(self.universe(self.corpus()), min_anchor=3)
        by_phrase = {p.text: c for c in clusters for p in c.members}
        self.assertIn("zigzag", by_phrase["quotex zigzag strategy"].label)

    def test_word_order_variants_collapse_into_one_keyword(self):
        universe = self.universe([
            "quotex ai trading bot", "ai quotex trading bot", "quotex trading bot ai",
            "quotex demo account", "quotex demo login", "quotex demo reset",
        ])
        clusters = clustering.build(universe, min_anchor=3)
        texts = [p.text for c in clusters for p in c.members]
        orderings = [t for t in texts if set(t.split()) == {"quotex", "ai", "trading", "bot"}]
        self.assertEqual(len(orderings), 1,
                         "three orderings of one keyword must not be three keywords")
        kept = [p for c in clusters for p in c.members if p.text in orderings][0]
        self.assertEqual(kept.variants, 3, "the collapsed forms must still be counted")

    def test_the_number_of_topics_is_bounded(self):
        """The old algorithm produced 1,796 clusters for 9,499 phrases; this cannot."""
        phrases = [f"quotex topic{i} thing" for i in range(60)] * 3
        phrases = [f"{p} {i}" for i, p in enumerate(phrases)]
        clusters = clustering.build(self.universe(phrases), topics=10, min_anchor=2)
        self.assertLessEqual(len(clusters), 11, "topics + the long tail")

    def test_a_keyword_with_no_topic_goes_to_the_long_tail_not_a_wrong_one(self):
        universe = self.universe(self.corpus() + ["quotex erfahrungen"])
        clusters = clustering.build(universe, min_anchor=3)
        home = next(c for c in clusters for p in c.members if p.text == "quotex erfahrungen")
        self.assertEqual(home.label, clustering.TAIL_LABEL,
                         "an unplaceable keyword must be named, not misfiled")

    def test_every_keyword_lands_in_exactly_one_cluster(self):
        phrases = self.corpus()
        clusters = clustering.build(self.universe(phrases), min_anchor=3)
        placed = [p.text for c in clusters for p in c.members]
        self.assertEqual(sorted(placed), sorted(phrases))
        self.assertEqual(len(placed), len(set(placed)))

    def test_clusters_come_back_in_priority_order(self):
        clusters = clustering.build(self.universe(self.corpus()), min_anchor=3)
        self.assertEqual([c.priority for c in clusters],
                         sorted([c.priority for c in clusters], reverse=True))


class IntentTests(unittest.TestCase):
    def classify(self, phrase):
        drop = clustering.STOPWORDS | set(seed_tokens("quotex"))
        return clustering.classify_intent(clustering.tokenize(phrase, drop))

    def test_each_intent_is_recognised(self):
        self.assertEqual(self.classify("quotex login"), "navigational")
        self.assertEqual(self.classify("quotex vs pocket option"), "commercial")
        self.assertEqual(self.classify("quotex bonus code"), "transactional")
        self.assertEqual(self.classify("how to use quotex"), "informational")
        self.assertEqual(self.classify("is quotex legal in india"), "informational")

    def test_an_unmarked_phrase_is_brand_not_navigational(self):
        self.assertEqual(self.classify("quotex signal bot"), clustering.BRAND)


class AskingForAMarket(unittest.TestCase):
    """`gl=` is a real dimension, and the only honest source of a market."""

    def test_the_market_reaches_the_query_string(self):
        client = SuggestClient(gl="ID")
        self.assertIn("gl=id", client.endpoint_url("binary option"))

    def test_no_market_asked_means_no_gl_sent(self):
        self.assertNotIn("gl=", SuggestClient().endpoint_url("x"))

    def test_a_market_is_part_of_the_cache_identity(self):
        cache = SuggestCache(None, egress="TR")
        plain = cache.key("chrome", "en", "", "q")
        self.assertNotEqual(plain, cache.key("chrome", "en", "", "q", "us"))
        self.assertNotEqual(cache.key("chrome", "en", "", "q", "us"),
                            cache.key("chrome", "en", "", "q", "id"))

    def test_responses_collected_before_markets_existed_still_replay(self):
        """The key must not change shape when no market is asked for."""
        cache = SuggestCache(None, egress="TR")
        self.assertEqual(cache.key("chrome", "en", "", "q"), "TR|chrome|en||q")
        self.assertEqual(cache.key("chrome", "en", "", "q", ""), "TR|chrome|en||q")


class MergingSeveralMarkets(unittest.TestCase):
    def _universe(self, seed, phrases):
        universe = Universe(seed=seed)
        for text, rank in phrases:
            phrase = Phrase(text, rank, 600, 0)
            phrase.parents.add(seed)
            universe.phrases[text] = phrase
        universe.exhausted = True
        return universe

    def test_a_keyword_records_every_market_that_returned_it(self):
        merged = merge_markets("binary option", {
            "US": self._universe("binary option", [("binary options cboe", 4),
                                                   ("binary options", 1)]),
            "ID": self._universe("binary option", [("binary options adalah", 3),
                                                   ("binary options", 2)]),
        })
        self.assertEqual(sorted(merged.phrases["binary options"].markets), ["ID", "US"])
        self.assertEqual(list(merged.phrases["binary options cboe"].markets), ["US"])
        self.assertEqual(list(merged.phrases["binary options adalah"].markets), ["ID"])

    def test_the_market_is_where_it_ranks_best_not_where_it_appeared_most(self):
        merged = merge_markets("x", {
            "US": self._universe("x", [("x thing", 9)]),
            "ID": self._universe("x", [("x thing", 2)]),
        })
        self.assertEqual(merged.phrases["x thing"].market, "ID")
        self.assertEqual(merged.phrases["x thing"].best_rank, 2)

    def test_the_markets_asked_are_named_on_the_run(self):
        merged = merge_markets("x", {"US": self._universe("x", [("x a", 1)]),
                                     "BR": self._universe("x", [("x b", 1)])})
        self.assertEqual(merged.market, "BR US")

    def test_one_market_still_holding_a_frontier_means_not_exhausted(self):
        open_market = self._universe("x", [("x a", 1)])
        open_market.exhausted = False
        merged = merge_markets("x", {"US": self._universe("x", [("x b", 1)]),
                                     "IN": open_market})
        self.assertFalse(merged.exhausted)

    def test_a_keyword_carries_no_market_when_none_was_asked(self):
        universe = self._universe("x", [("x a", 1)])
        self.assertEqual(universe.phrases["x a"].market, "")
        self.assertEqual(universe.phrases["x a"].markets, {})


class TheExitCountryIsNeverAMarket(unittest.TestCase):
    """The defect this replaced: a country read off whichever proxy answered."""

    def test_a_phrase_row_offers_a_market_and_never_an_exit_country(self):
        phrase = Phrase("x a", 1, 600, 0)
        phrase.markets = {"US": 1}
        row = phrase.as_row()
        self.assertIn("market", row)
        self.assertIn("markets", row)
        self.assertNotIn("country", row)
        self.assertNotIn("countries", row)

    def test_the_exit_tally_is_kept_on_the_run_where_it_cannot_mislead(self):
        universe = Universe(seed="x")
        self.assertEqual(universe.egress_countries, {})


class TheProxyLimitsAreNotRestatedHere(unittest.TestCase):
    """Two sources of truth for one limit is how 0.2/10/200 outlived itself."""

    def test_this_package_holds_no_copy_of_the_numbers(self):
        from keel_seo.keywords import proxying

        if proxying.AVAILABLE:
            from keel_crawler.proxy.pool import PER_PROXY_RPS as owner
            self.assertEqual(proxying.PER_PROXY_RPS, owner)
        else:
            self.assertIsNone(proxying.PER_PROXY_RPS)
            self.assertIsNone(proxying.PER_PROXY_PER_MINUTE)
            self.assertIsNone(proxying.PER_PROXY_PER_HOUR)


class TheHarvestWalker(unittest.TestCase):
    """The service layer, which every project now shares instead of copying."""

    def test_comments_and_blank_lines_are_not_seeds(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "seeds.txt")
            with open(path, "w", encoding="utf-8") as handle:
                handle.write("# a comment\n\nquotex\n  pocket option  # inline\n\n")
            self.assertEqual(harvest.read_seeds(path), ["quotex", "pocket option"])

    def test_a_seed_with_output_on_disk_is_considered_done(self):
        with tempfile.TemporaryDirectory() as tmp:
            self.assertFalse(harvest.already_harvested(tmp, "pocket option"))
            for kind in ("json", "md"):
                Path(tmp, f"pocket-option.{kind}").write_text("{}", encoding="utf-8")
            self.assertTrue(harvest.already_harvested(tmp, "pocket option"))

    def test_the_walker_looks_for_the_same_filename_the_writer_writes(self):
        """Two spellings of one stem is how a walker re-pays for finished work."""
        self.assertEqual(report.slugify("pocket option"), "pocket-option")
        self.assertEqual(report.slugify("Quotex"), "quotex")
        self.assertEqual(report.slugify("gold & silver"), "gold-silver")

    def test_an_empty_seed_list_is_success_and_not_a_failure(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "seeds.txt")
            Path(path).write_text("# nothing but a comment\n", encoding="utf-8")
            code = harvest.main(["--seeds", path, "--out", os.path.join(tmp, "out")])
            self.assertEqual(code, 0)

    def test_every_seed_already_done_makes_the_walk_free(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out")
            os.makedirs(out)
            for kind in ("json", "md"):
                Path(out, f"quotex.{kind}").write_text("{}", encoding="utf-8")
            path = os.path.join(tmp, "seeds.txt")
            Path(path).write_text("quotex\n", encoding="utf-8")
            code = harvest.main(["--seeds", path, "--out", out])
            self.assertEqual(code, 0)


class ThePackageVersion(unittest.TestCase):
    """It was reporting 0.26.0 from a 0.27.0 install, and nothing failed."""

    def test_it_matches_the_one_file_ci_actually_guards(self):
        import re

        import keel_seo

        pyproject = Path(__file__).resolve().parents[1] / "pyproject.toml"
        declared = re.search(r'^version\s*=\s*"([^"]+)"',
                             pyproject.read_text(encoding="utf-8"), re.M)
        self.assertIsNotNone(declared)
        self.assertEqual(keel_seo.__version__, declared.group(1))


class TheSyncFileList(unittest.TestCase):
    def test_the_response_cache_is_not_a_harvest_format(self):
        self.assertNotIn("jsonl", sync.FORMATS)
        self.assertEqual(set(sync.FORMATS), {"xlsx", "json", "csv", "md"})


class OneSeedManySpellings(unittest.TestCase):
    """`fundingpips` and `funding pips` are one brand, so they are one harvest.

    Before this, a seed matched against the spaced phrase recognised only its own
    spelling: a `fundingpips` run filed every `funding pips ...` suggestion as
    contamination, and the two spellings of one keyword — when both did get in —
    described different topics and landed in different clusters.
    """

    def test_the_spaced_spelling_is_no_longer_contamination(self):
        tokens = seed_tokens("fundingpips")
        self.assertTrue(contains_seed("funding pips rules", tokens))
        self.assertTrue(contains_seed("funding-pips payout", tokens))
        self.assertTrue(contains_seed("fundingpips rules", tokens))
        self.assertFalse(contains_seed("funded next rules", tokens))

    def test_a_multi_word_seed_still_matches_in_any_order(self):
        tokens = seed_tokens("pip value calculator")
        self.assertTrue(contains_seed("calculator for pip value", tokens))
        self.assertFalse(contains_seed("pip calculator", tokens))

    def test_squash_removes_every_separator(self):
        self.assertEqual(squash("Funding-Pips  Rules"), "fundingpipsrules")

    def test_the_spelling_a_phrase_uses_is_read_back_from_it(self):
        tokens = seed_tokens("fundingpips")
        self.assertEqual(seed_spelling("funding pips rules", tokens), "funding pips")
        self.assertEqual(seed_spelling("best fundingpips deal", tokens), "fundingpips")
        # A multi-word seed has no single spelling to extract, and guessing one
        # would name a span that is not the seed.
        self.assertEqual(seed_spelling("calculator for pip value",
                                       seed_tokens("pip value calculator")), "")

    def test_only_spellings_google_keeps_returning_are_chased(self):
        universe = Universe(seed="fundingpips")
        for text in ("funding pips review", "funding pips rules",
                     "funding pips payout", "fundingpips login"):
            universe.phrases[text] = Phrase(text, 1, 600, 0)
        universe.phrases["funding-pips x"] = Phrase("funding-pips x", 1, 600, 0)

        found = discover_variants(universe, seed_tokens("fundingpips"),
                                  "fundingpips", limit=2)
        self.assertEqual(found, ["funding pips"],
                         "one sighting of 'funding-pips' is a typo, not a spelling")

    def test_both_spellings_of_one_keyword_share_one_cluster(self):
        universe = Universe(seed="fundingpips", variants=("funding pips",))
        for text in ("fundingpips rules", "funding pips rules",
                     "fundingpips payout", "funding pips payout",
                     "fundingpips payout speed"):
            universe.phrases[text] = Phrase(text, 1, 600, 0)
        score(universe)

        clusters = clustering.build(universe, topics=5, min_anchor=1)
        placed = {p.text: c.index for c in clusters for p in c.members}
        collapsed = {p.text: p.also_written for c in clusters for p in c.members}

        self.assertEqual(len(placed), 3, "each keyword should survive once")
        self.assertIn("rules", [c.label for c in clusters])
        # The two spellings of "rules" are one keyword, and the row says so.
        head = next(t for t in collapsed if t.endswith("rules"))
        self.assertEqual(collapsed[head], [
            "funding pips rules" if head == "fundingpips rules" else "fundingpips rules"
        ])

    def test_without_the_variant_the_spellings_would_split(self):
        """The witness for the fix: the same phrases, with no variant declared."""
        universe = Universe(seed="fundingpips")
        for text in ("fundingpips rules", "funding pips rules"):
            universe.phrases[text] = Phrase(text, 1, 600, 0)
        score(universe)

        clusters = clustering.build(universe, topics=5, min_anchor=1)
        keywords = [p.text for c in clusters for p in c.members]
        self.assertEqual(len(keywords), 2,
                         "unknown spelling: they stay two keywords, which is the "
                         "behaviour the variant list exists to fix")

    def test_markets_merge_the_spellings_each_of_them_found(self):
        us = Universe(seed="fundingpips", variants=("funding pips",))
        us.phrases["fundingpips rules"] = Phrase("fundingpips rules", 1, 600, 0)
        de = Universe(seed="fundingpips", variants=("funding-pips",))
        de.phrases["fundingpips regeln"] = Phrase("fundingpips regeln", 1, 600, 0)

        merged = merge_markets("fundingpips", {"US": us, "DE": de})
        self.assertEqual(sorted(merged.variants), ["funding pips", "funding-pips"])


class ProbingBeforeBuying(unittest.TestCase):
    """A market is sampled before it is crawled, and most of them are set aside.

    Sixteen markets is sixteen crawls, and most secondary markets return the
    primary market's own answers in a different accent. The probe asks each one a
    sixtieth of a seed tier and keeps only the markets that answer differently.
    """

    def test_a_market_that_echoes_the_primary_is_set_aside(self):
        shared = [f"ftmo topic {n}" for n in range(12)]
        client = _AlwaysAnswers(shared)
        _, verdict = probe_market("ftmo", client, set(shared), queries=10)
        self.assertEqual(verdict["new"], 0)
        self.assertEqual(verdict["novelty"], 0.0)
        self.assertFalse(worth_crawling(verdict))

    def test_a_market_that_answers_differently_earns_its_crawl(self):
        local = [f"ftmo erfahrungen {n}" for n in range(40)]
        client = _AlwaysAnswers(local)
        universe, verdict = probe_market("ftmo", client, {"ftmo review"}, queries=10)
        self.assertEqual(verdict["new"], 40)
        self.assertEqual(verdict["novelty"], 1.0)
        self.assertTrue(worth_crawling(verdict))
        self.assertEqual(len(universe.phrases), 40,
                         "the probe's own findings are kept, not thrown away")

    def test_the_threshold_separates_the_markets_it_was_measured_on(self):
        """The measured distribution, kept where a future edit has to face it.

        Probing all fifteen secondary target markets against a US primary (seed
        `fundingpips`, 60 queries each, 2026-09-04). The two groups are the
        markets asked in another language and the markets asked in English, and
        nothing lands between 17% and 29%.
        """
        measured = {
            "ES": 0.42, "ID": 0.40, "AR": 0.40, "DE": 0.32, "FR": 0.30,
            "PT": 0.29, "BR": 0.29,
            "IN": 0.17, "NG": 0.12, "KE": 0.12, "PK": 0.11,
            "CA": 0.05, "ZA": 0.05, "PH": 0.05, "MY": 0.05,
        }
        different = {"ES", "ID", "AR", "DE", "FR", "PT", "BR"}
        for code, novelty in measured.items():
            verdict = {"new": 100, "novelty": novelty}
            self.assertEqual(worth_crawling(verdict), code in different,
                             f"{code} at {novelty:.0%} falls on the wrong side")

    def test_both_tests_must_pass_not_either(self):
        # All new, but four phrases: 100% of nothing.
        self.assertFalse(worth_crawling({"new": 4, "novelty": 1.0}))
        # Plenty new, but a rounding error of the whole: busy agreeing.
        self.assertFalse(worth_crawling({"new": 30, "novelty": 0.03}))
        self.assertTrue(worth_crawling({"new": 30, "novelty": 0.4}))

    def test_the_sample_spans_the_attachment_families(self):
        """Taking the first N would ask nothing but a-z suffixes."""
        client = _AlwaysAnswers([])
        probe_market("quotex", client, set(), queries=60)
        asked = client.asked
        self.assertEqual(len(asked), 60)
        self.assertTrue(any(q.startswith("quotex ") for q in asked), "no suffix sweep")
        self.assertTrue(any(q.endswith(" quotex") for q in asked), "no prefix sweep")
        self.assertTrue(any(q.startswith("quotex") and " " not in q for q in asked),
                        "no tight sweep")

    def test_a_probe_asks_a_thousandth_of_what_a_crawl_would(self):
        """The number that makes the question worth asking at all."""
        seed_tier = len(expansions("quotex", SEED))
        self.assertLess(PROBE_QUERIES, seed_tier / 5)

    def test_the_cli_offers_the_knobs_and_defaults_to_probing(self):
        from keel_seo.keywords.__main__ import build_parser

        args = build_parser().parse_args(["quotex"])
        self.assertEqual(args.probe, PROBE_QUERIES)
        self.assertEqual(args.probe_share, PROBE_NOVELTY_SHARE)
        self.assertEqual(args.probe_floor, PROBE_NOVELTY_FLOOR)
        self.assertEqual(args.primary, "", "the primary defaults to the list head")


class _AlwaysAnswers(SuggestClient):
    """Answers every query with the same phrases, and remembers what was asked."""

    def __init__(self, phrases):
        super().__init__(cache=SuggestCache(None))
        self.phrases = list(phrases)
        self.asked: list[str] = []

    def fetch(self, query):
        self.asked.append(query)
        return Response(
            query=query,
            suggestions=tuple(Suggestion(p, i, 600 - i)
                              for i, p in enumerate(self.phrases, 1)),
            capacity=15,
        )


class TheTargetMarkets(unittest.TestCase):
    """Sixteen countries, each asked in the language it searches in."""

    def test_the_default_list_is_the_sixteen_target_countries(self):
        self.assertEqual(list(target_markets.TARGET_MARKETS), [
            "US", "CA", "DE", "FR", "ES", "PT", "BR", "AR",
            "IN", "PK", "ZA", "NG", "KE", "PH", "MY", "ID"])

    def test_a_market_is_asked_in_the_language_it_searches_in(self):
        self.assertEqual(target_markets.language_for("BR"), "pt")
        self.assertEqual(target_markets.language_for("DE"), "de")
        self.assertEqual(target_markets.language_for("ID"), "id")
        # English is not a fallback here, it is the answer: these markets search
        # in English, and asking them in a local language returns a smaller and
        # less commercial universe than the one their searchers use.
        for code in ("IN", "PK", "NG", "KE", "ZA", "PH", "MY"):
            self.assertEqual(target_markets.language_for(code), "en", code)

    def test_an_explicit_language_wins_everywhere(self):
        self.assertEqual(target_markets.language_for("BR", "en"), "en")

    def test_a_list_can_be_written_any_way_a_person_would_write_it(self):
        self.assertEqual(target_markets.parse("us, in br"), ["US", "IN", "BR"])
        self.assertEqual(target_markets.parse("us,us,in"), ["US", "IN"])
        self.assertEqual(target_markets.parse("target"),
                         list(target_markets.TARGET_MARKETS))
        self.assertEqual(target_markets.parse(""), [])

    def test_a_country_name_is_not_a_country_code(self):
        with self.assertRaises(target_markets.UnknownMarket):
            target_markets.parse("usa")

    def test_the_project_can_override_the_default(self):
        os.environ[target_markets.ENV_NAME] = "us,de"
        try:
            self.assertEqual(target_markets.resolve(), ["US", "DE"])
            self.assertEqual(target_markets.resolve("br"), ["BR"],
                             "the command line beats the project setting")
        finally:
            del os.environ[target_markets.ENV_NAME]
        self.assertEqual(target_markets.resolve(), list(target_markets.TARGET_MARKETS))

    def test_the_cli_defers_the_default_instead_of_owning_it(self):
        from keel_seo.keywords.__main__ import build_parser

        self.assertIsNone(build_parser().parse_args(["quotex"]).markets,
                          "a default here would shadow the project's setting")


class TheLanguageColumn(unittest.TestCase):
    """Which language, not merely "not English" — and with no model."""

    def test_script_names_the_language_outright(self):
        self.assertEqual(language.language_of("ftmo क्या है"), "Hindi")
        self.assertEqual(language.language_of("ftmo отзывы"), "Russian")
        self.assertEqual(language.language_of("ftmo 후기"), "Korean")

    def test_vocabulary_names_it_when_the_letters_cannot(self):
        self.assertEqual(language.language_of("ftmo erfahrungen"), "German")
        self.assertEqual(language.language_of("ftmo opiniones"), "Spanish")
        self.assertEqual(language.language_of("ftmo corretora saque"), "Portuguese")
        self.assertEqual(language.language_of("ftmo cara daftar"), "Indonesian")

    def test_a_telling_letter_names_it_when_no_word_does(self):
        self.assertEqual(language.language_of("ftmo español"), "Spanish")
        self.assertEqual(language.language_of("ftmo avaliação"), "Portuguese")

    def test_a_shared_accent_is_reported_as_unnamed_rather_than_guessed(self):
        language_name, why = language.identify("ftmo café")
        self.assertEqual(language_name, language.UNDETERMINED)
        self.assertEqual(why, language.DIACRITIC)

    def test_english_stays_english(self):
        for text in ("ftmo review", "quotex com login", "quotex dao download",
                     "fundingpips payout speed"):
            self.assertEqual(language.language_of(text), language.ENGLISH, text)
            self.assertEqual(language.non_english_reason(text), "")

    def test_the_evidence_travels_with_the_verdict(self):
        language_name, why = language.identify("ftmo erfahrungen")
        self.assertEqual(language_name, "German")
        self.assertIn("erfahrungen", why)


class TheEgressRule(unittest.TestCase):
    """A harvest never leaves from the machine running it.

    Measured 2026-09-04: a run at the throttled default of 6 q/s, asking from one
    production address, was refused after 3,909 requests and left its seed two
    thirds unexpanded. The block is IP-wide, so that host lost the endpoint for
    everything it asked, not only for the crawl. These tests exist because the
    two entry points had already drifted apart once — the batch walker defaulted
    to rotation for months while the single-seed CLI still defaulted to direct,
    and the harvest that earned the block went through the second one.
    """

    def test_both_entry_points_default_to_rotation(self):
        from keel_seo.keywords.__main__ import build_parser as crawler_parser

        self.assertEqual(crawler_parser().parse_args(["quotex"]).proxies, "auto")
        self.assertEqual(
            harvest.build_parser().parse_args(["--seeds", "s", "--out", "o"]).proxies,
            "auto")

    def test_every_spelling_of_direct_egress_is_refused(self):
        for mode in ("off", "OFF", "none", "direct", ""):
            with self.subTest(mode=mode):
                with self.assertRaises(DirectEgressRefused):
                    require_pooled_egress(mode)

    def test_rotation_is_accepted_and_returned_unchanged(self):
        self.assertEqual(require_pooled_egress("auto"), "auto")

    def test_the_crawler_refuses_before_it_creates_anything(self):
        """A run that may not ask must not leave an output directory behind."""
        from keel_seo.keywords.__main__ import main as crawler_main

        with tempfile.TemporaryDirectory() as tmp:
            out = os.path.join(tmp, "out")
            code = crawler_main(["quotex", "--out", out, "--proxies", "off"])
            self.assertEqual(code, 1)
            self.assertFalse(os.path.exists(out),
                             "the refusal came after the directory was made")

    def test_the_walker_refuses_once_rather_than_per_seed(self):
        with tempfile.TemporaryDirectory() as tmp:
            seeds = os.path.join(tmp, "seeds.txt")
            Path(seeds).write_text("quotex\nftmo\n", encoding="utf-8")
            out = os.path.join(tmp, "out")
            code = harvest.main(["--seeds", seeds, "--out", out, "--proxies", "off"])
            self.assertEqual(code, 1)
            self.assertFalse(os.path.exists(out))

    def test_the_refusal_says_what_to_do_instead(self):
        """An error that only forbids sends the reader hunting for a flag to flip."""
        self.assertIn("--proxies auto", DIRECT_REFUSAL)
        self.assertIn("3,909", DIRECT_REFUSAL)


if __name__ == "__main__":
    unittest.main()
