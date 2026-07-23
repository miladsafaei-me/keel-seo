#!/usr/bin/env python3
"""Persistent query registry — the durable GSC query store.

The registry is the durable record of every Search Console query a property has
ever appeared for, plus an optional per-query ``cluster_id`` a host's clustering
step fills in. Each ``sync`` pulls the full query universe from the API and merges
it in: existing queries get refreshed metrics and a bumped ``last_seen``; genuinely
new queries land with ``cluster_id=null`` and ``status="new"`` so an incremental
clustering step knows exactly what to look at.

Storage (``$GSC_DATA_DIR``, default ``~/.local/share/keel-seo/gsc``):
    registry.json            the authoritative store (query -> record)
    registry.csv             human/skill-readable view, sorted by clicks
    snapshots/YYYY-MM-DD.json raw pull kept for audit

Run from a venv with the ``[gsc]`` extra installed:
    python -m keel_seo.gsc.registry sync
    python -m keel_seo.gsc.registry stats
    python -m keel_seo.gsc.registry xlsx
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import sys
from pathlib import Path

from . import connector as gsc

DEFAULT_DATA_DIR = Path.home() / ".local" / "share" / "keel-seo" / "gsc"
DATA_LAG_DAYS = 2
FULL_WINDOW_DAYS = 480  # ~16 months = the GSC domain-property max; captures the whole universe


def _data_dir() -> Path:
    d = Path(os.environ.get("GSC_DATA_DIR", str(DEFAULT_DATA_DIR))).expanduser()
    (d / "snapshots").mkdir(parents=True, exist_ok=True)
    return d


def _registry_path() -> Path:
    return _data_dir() / "registry.json"


def _load_registry() -> dict:
    path = _registry_path()
    if path.exists():
        return json.loads(path.read_text())
    return {"site": gsc.DEFAULT_SITE, "updated": "", "queries": {}}


def _pull_full(site: str, days: int) -> "tuple[list[dict], str, str]":
    service, _ = gsc._load_service()
    end = dt.date.today() - dt.timedelta(days=DATA_LAG_DAYS)
    start = end - dt.timedelta(days=days - 1)
    body = {
        "startDate": start.isoformat(),
        "endDate": end.isoformat(),
        "dimensions": ["query"],
        "rowLimit": 25000,
    }
    rows: list[dict] = []
    start_row = 0
    while True:  # paginate defensively even though the universe fits one page today
        body["startRow"] = start_row
        try:
            resp = service.searchanalytics().query(siteUrl=site, body=body).execute()
        except Exception as exc:
            gsc._fail(gsc._explain_api_error(exc))
        page = resp.get("rows", [])
        rows.extend(page)
        if len(page) < 25000:
            break
        start_row += 25000
    return rows, start.isoformat(), end.isoformat()


def cmd_sync(args) -> None:
    site = gsc._require_site(args.site)
    rows, start, end = _pull_full(site, args.days)
    today = dt.date.today().isoformat()

    snapshot = _data_dir() / "snapshots" / f"{today}.json"
    snapshot.write_text(json.dumps({"site": site, "start": start, "end": end, "rows": rows}, indent=2))

    reg = _load_registry()
    queries = reg["queries"]
    added = 0
    for r in rows:
        keys = r.get("keys", [])
        if not keys:
            continue
        q = keys[0]
        metrics = {
            "clicks": int(r.get("clicks", 0)),
            "impressions": int(r.get("impressions", 0)),
            "ctr": round(r.get("ctr", 0.0), 4),
            "position": round(r.get("position", 0.0), 1),
        }
        rec = queries.get(q)
        if rec is None:
            queries[q] = {
                **metrics,
                "first_seen": today,
                "last_seen": today,
                "cluster_id": None,
                "status": "new",
            }
            added += 1
        else:
            rec.update(metrics)
            rec["last_seen"] = today

    reg["site"] = site
    reg["updated"] = today
    _registry_path().write_text(json.dumps(reg, indent=2))
    _write_csv(reg)

    unclustered = sum(1 for v in queries.values() if not v["cluster_id"])
    print(f"synced {site}  window {start} -> {end}")
    print(f"  pulled     : {len(rows)} queries this window")
    print(f"  registry   : {len(queries)} total unique queries")
    print(f"  new        : {added} added this run (status=new)")
    print(f"  unclustered: {unclustered} queries have no cluster yet")
    print(f"  snapshot   : {snapshot}")


def _write_csv(reg: dict) -> None:
    rows = sorted(reg["queries"].items(), key=lambda kv: kv[1]["clicks"], reverse=True)
    path = _data_dir() / "registry.csv"
    with open(path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(
            ["query", "clicks", "impressions", "ctr", "position",
             "first_seen", "last_seen", "cluster_id", "status"]
        )
        for q, v in rows:
            w.writerow([
                q, v["clicks"], v["impressions"], v["ctr"], v["position"],
                v["first_seen"], v["last_seen"], v["cluster_id"] or "", v["status"],
            ])


def cmd_xlsx(args) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    reg = _load_registry()
    if not reg["queries"]:
        gsc._fail("registry is empty — run `sync` first.")
    rows = sorted(reg["queries"].items(), key=lambda kv: kv[1]["clicks"], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Queries"
    headers = ["#", "query", "clicks", "impressions", "ctr %", "position",
               "first_seen", "last_seen", "cluster_id", "status"]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F5B3A")
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = head_font
        ws.cell(row=1, column=c).fill = head_fill

    for i, (q, v) in enumerate(rows, start=1):
        ws.append([
            i, q, v["clicks"], v["impressions"], round(v["ctr"] * 100, 2), v["position"],
            v["first_seen"], v["last_seen"], v["cluster_id"] or "", v["status"],
        ])

    widths = [5, 46, 9, 12, 8, 9, 12, 12, 16, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    out = Path(args.out) if args.out else (_data_dir() / "query-registry.xlsx")
    wb.save(out)
    print(f"wrote {out}  ({len(rows)} queries)")


def cmd_stats(_args) -> None:
    reg = _load_registry()
    q = reg["queries"]
    if not q:
        print("registry is empty — run `sync` first.")
        return
    unclustered = sum(1 for v in q.values() if not v["cluster_id"])
    new = sum(1 for v in q.values() if v["status"] == "new")
    total_clicks = sum(v["clicks"] for v in q.values())
    total_impr = sum(v["impressions"] for v in q.values())
    print(f"registry: {reg['site']}  (updated {reg['updated']})")
    print(f"  total queries : {len(q)}")
    print(f"  unclustered   : {unclustered}")
    print(f"  status=new    : {new}")
    print(f"  total clicks  : {total_clicks}")
    print(f"  total impress : {total_impr}")


def main() -> None:
    p = argparse.ArgumentParser(prog="keel-seo-gsc-registry", description="GSC query registry (pull + merge)")
    sub = p.add_subparsers(dest="command", required=True)

    s = sub.add_parser("sync", help="pull the full query universe and merge into the registry")
    s.add_argument("--site", default=gsc.DEFAULT_SITE)
    s.add_argument("--days", type=int, default=FULL_WINDOW_DAYS)
    s.set_defaults(func=cmd_sync)

    x = sub.add_parser("xlsx", help="export the registry to a formatted .xlsx")
    x.add_argument("--out", help="output path (default $GSC_DATA_DIR/query-registry.xlsx)")
    x.set_defaults(func=cmd_xlsx)

    sub.add_parser("stats", help="print registry summary").set_defaults(func=cmd_stats)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
