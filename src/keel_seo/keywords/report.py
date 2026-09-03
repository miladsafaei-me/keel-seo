"""Write a crawl out four ways, for four different readers.

JSON is the record a later run or another script reads. CSV is the flat form.
Markdown is what a person reads to decide what to build. XLSX is the one people
actually work in — sheets for the keywords, the clusters, the provenance and the
contamination check, with filters already set up.

All of them carry the run's metadata, and the metadata always names the egress
country rather than a requested one, because that is the only geography an
autocomplete harvest actually has. The volume caveat travels inside every format
too: a file outlives the terminal that produced it.
"""
from __future__ import annotations

import csv
import datetime as dt
import json
import os

from .cluster import Cluster
from .crawl import Universe

CONTAMINATION_SAMPLE = 40


def metadata(universe: Universe, clusters: list[Cluster], egress: dict,
             client) -> dict:
    pool = getattr(client, "pool", None)
    # Clustering collapses word-order variants and plurals, so the number of
    # KEYWORDS is smaller than the number of raw phrases returned. Reporting the
    # raw count next to a sheet holding the collapsed one is a contradiction the
    # reader has to resolve, so both are named and the difference is stated.
    keywords = sum(len(c.members) for c in clusters)
    return {
        "seed": universe.seed,
        "harvested_at": dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": "google-autocomplete",
        "endpoint_client": client.client,
        "language": client.hl,
        "vertical": client.ds or "web",
        "egress_ip": egress.get("ip", ""),
        "egress_country": egress.get("country", "unknown"),
        "egress_org": egress.get("org", ""),
        "phrases": len(universe.phrases),
        "keywords": keywords,
        "variants_collapsed": len(universe.phrases) - keywords,
        "clusters": len(clusters),
        "queries_asked": universe.queries_asked,
        "network_calls": universe.network_calls,
        "cache_hits": universe.cache_hits,
        "errors": universe.errors,
        "levels_run": universe.levels_run,
        "unexpanded_phrases": universe.unexpanded,
        "exhausted": universe.exhausted,
        "stopped_by_rate_limit": universe.blocked,
        "stopped_by_time_limit": universe.timed_out,
        "rate_limited_responses": universe.rate_limited,
        "elapsed_seconds": round(universe.elapsed, 1),
        "off_seed_rejected": len(universe.off_seed),
        "proxy_pool": pool.stats() if pool is not None else None,
        "per_level": universe.per_level,
        "volume_note": (
            "Autocomplete never returns search volume. priority ranks demand "
            "shape (Google's own ordering, breadth of reach, relevance score, "
            "depth) and is not a volume estimate."
        ),
    }


def write_json(path: str, universe: Universe, clusters: list[Cluster],
               meta: dict) -> None:
    contamination = sorted(universe.off_seed.items(), key=lambda kv: -kv[1])
    payload = {
        "meta": meta,
        "clusters": [c.as_row() for c in clusters],
        "off_seed_sample": [
            {"phrase": phrase, "times_returned": count}
            for phrase, count in contamination[:CONTAMINATION_SAMPLE]
        ],
    }
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=1)


def write_csv(path: str, clusters: list[Cluster]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            ["priority", "keyword", "cluster", "cluster_label", "intent",
             "best_rank", "reach", "relevance", "level", "words", "variants"]
        )
        for cluster in clusters:
            for phrase in cluster.members:
                writer.writerow(
                    [round(phrase.priority, 1), phrase.text, cluster.index,
                     cluster.label, cluster.intent, phrase.best_rank, phrase.reach,
                     phrase.max_relevance, phrase.first_level, phrase.words,
                     phrase.variants]
                )


def write_markdown(path: str, universe: Universe, clusters: list[Cluster],
                   meta: dict, *, per_cluster: int = 12) -> None:
    lines: list[str] = []
    add = lines.append
    add(f"# Keyword universe — `{universe.seed}`")
    add("")
    if meta.get("egress_country") == "mixed":
        add(f"**Source:** Google autocomplete only ({meta['endpoint_client']} client, "
            f"`hl={meta['language']}`, {meta['vertical']} vertical). "
            f"**Egress: mixed** — {meta['egress_org']}. Autocomplete answers by "
            "requesting IP, so a rotating pool produces a deliberately "
            "multi-country harvest. Read this as *what the term is asked, broadly* "
            "rather than as one market's demand: some phrases below will be local "
            "to a single country, and the exact mix is not reproducible.")
    else:
        add(f"**Source:** Google autocomplete only ({meta['endpoint_client']} client, "
            f"`hl={meta['language']}`, {meta['vertical']} vertical). "
            f"**Egress:** {meta['egress_country']} ({meta['egress_ip']}) — autocomplete "
            "geography is the requesting IP, never a parameter, so this is the market "
            "the harvest actually reflects.")
    add("")
    add(f"**Harvested:** {meta['harvested_at']} · "
        f"**{meta.get('keywords', meta['phrases']):,} keywords** in "
        f"**{meta['clusters']} clusters** from {meta['queries_asked']:,} queries in "
        f"{meta['elapsed_seconds']}s"
        + (f" ({meta['variants_collapsed']:,} re-worded duplicates collapsed)."
           if meta.get("variants_collapsed") else "."))
    add("")
    add("> Autocomplete returns no search volume, and no parameter exists that "
        "would make it. `priority` below ranks demand *shape* — Google's own "
        "ordering, how many independent expansions surfaced a phrase, its "
        "relevance score and its depth. It is not a volume estimate and must not "
        "be presented as one.")
    add("")
    if universe.blocked:
        add(f"⚠️ **Google rate-limited this harvest** after "
            f"{meta['network_calls']:,} requests and it stopped early, with "
            f"{meta['unexpanded_phrases']:,} phrases left unexpanded. What is below "
            "was collected before the block and is sound; it is not the complete "
            "universe. Re-run later with a lower `--rate` — the cache means the "
            "work already done is not repeated.")
        add("")
    elif universe.timed_out:
        add(f"⏱️ **The crawl reached its time limit** and stopped cleanly, with "
            f"{meta['unexpanded_phrases']:,} phrases still unexpanded. Everything "
            "below was collected and kept — nothing was lost to the deadline. Every "
            "response is cached, so re-running resumes here rather than restarting.")
        add("")
    elif not universe.exhausted:
        add(f"⚠️ The crawl stopped with **{meta['unexpanded_phrases']:,} phrases "
            "left unexpanded** (budget or level cap reached), so this universe is "
            "wide but not closed. Re-run with a higher `--budget` / `--levels` to "
            "continue; the cache makes the repeat work free.")
        add("")
    else:
        add("✅ The crawl closed on its own: every phrase it found was re-seeded "
            "and produced nothing new. This is the complete universe at this "
            "grammar, language and egress.")
        add("")

    add("## Clusters, most valuable first")
    add("")
    add("| # | Cluster | Intent | Phrases | Priority | Head phrase |")
    add("|---|---|---|---|---|---|")
    for cluster in clusters:
        add(f"| {cluster.index} | {cluster.label} | {cluster.intent} | "
            f"{cluster.size} | {cluster.priority:.0f} | {cluster.head.text} |")
    add("")

    add("## Inside each cluster")
    add("")
    for cluster in clusters:
        add(f"### {cluster.index}. {cluster.label} "
            f"({cluster.intent}, {cluster.size} phrases)")
        add("")
        add("| Priority | Keyword | Rank | Reach | Level |")
        add("|---|---|---|---|---|")
        for phrase in cluster.members[:per_cluster]:
            add(f"| {phrase.priority:.0f} | {phrase.text} | {phrase.best_rank} | "
                f"{phrase.reach} | {phrase.first_level} |")
        if cluster.size > per_cluster:
            add(f"| | *…{cluster.size - per_cluster} more in the CSV* | | | |")
        add("")

    contamination = sorted(universe.off_seed.items(), key=lambda kv: -kv[1])
    if contamination:
        add("## Off-seed neighbours (contamination check)")
        add("")
        add("Phrases Google returned that do **not** contain the seed. A term "
            "whose neighbours belong to another industry is a term whose traffic "
            "will too.")
        add("")
        add("| Times returned | Phrase |")
        add("|---|---|")
        for phrase, count in contamination[:CONTAMINATION_SAMPLE]:
            add(f"| {count} | {phrase} |")
        add("")

    with open(path, "w", encoding="utf-8") as handle:
        handle.write("\n".join(lines) + "\n")


# Short enough to be read standing up. The reasoning behind the weights, the
# measured correlations and the tuning evidence live in the package README - a
# glossary nobody finishes explains nothing, and this one sits in a spreadsheet
# where every extra clause is a row the reader scrolls past.
COLUMN_MEANINGS = (
    ("Priority", "How promising the keyword is, 0-100. Sort by this. It measures how "
                 "much demand a keyword shows, not how many searches it gets - "
                 "Google's autocomplete never reveals search volume."),
    ("Reach", "How many different Google queries returned this phrase. The best "
              "signal in the file: what many queries surface sits at the centre of "
              "the topic."),
    ("Relevance", "Google's own score for the suggestion. Higher is stronger."),
    ("Level", "Which round found it. 0 came straight from the seed; higher numbers "
              "are deeper in the tail."),
    ("Best rank", "Its best position in Google's suggestion list. 1 is the top."),
    ("Variants", "How many ways the same keyword was typed, merged into this one "
                 "row. 1 means it was only seen one way."),
    ("Intent", "What the searcher wants: reach a page (navigational), learn "
               "something (informational), compare before buying (commercial), act "
               "now (transactional), or simply the brand."),
)


def write_xlsx(path: str, universe: Universe, clusters: list[Cluster],
               meta: dict) -> bool:
    """Write the workbook people actually work in. False if openpyxl is absent.

    Four sheets, in the order they are used. **Run** comes first because it is
    what a reader needs before trusting any number in the file: where the data
    came from, how complete it is, and what each column means. **Clusters** is
    the map - one row per topic, the unit a page is built against - and every row
    links straight to its own keywords. **Keywords** is the working sheet, sorted
    by priority within cluster, with a frozen header and an autofilter.
    **Off-seed** is the contamination check.

    The Run sheet deliberately carries a short summary rather than every field in
    the metadata: proxy statistics and internal counters belong in the JSON, and
    a summary nobody reads is the same as no summary. The volume caveat is not a
    row of its own either - it lives inside the explanation of Priority, which is
    the number it applies to and the place a reader is actually looking.

    openpyxl is an optional dependency (`pip install 'keel-seo[xlsx]'`); the other
    three formats are stdlib and always written, so a missing library costs the
    workbook and nothing else.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.hyperlink import Hyperlink
    except ImportError:
        return False

    book = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    title_font = Font(bold=True, size=12, color="1F3864")
    link_font = Font(color="0563C1", underline="single")

    def style_header(sheet, widths, row=1):
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for cell in sheet[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

    run = book.active
    run.title = "Run"
    clusters_sheet = book.create_sheet("Clusters")
    keywords = book.create_sheet("Keywords")

    keywords.append(["Priority", "Keyword", "Cluster", "Cluster label", "Intent",
                     "Best rank", "Reach", "Relevance", "Level", "Words", "Variants"])
    # Remember where each cluster starts so the Clusters sheet can link to it.
    first_row: dict[int, int] = {}
    for cluster in clusters:
        first_row[cluster.index] = keywords.max_row + 1
        for phrase in cluster.members:
            keywords.append([round(phrase.priority, 1), phrase.text, cluster.index,
                             cluster.label, cluster.intent, phrase.best_rank,
                             phrase.reach, phrase.max_relevance, phrase.first_level,
                             phrase.words, phrase.variants])
    style_header(keywords, [9, 52, 9, 30, 15, 11, 8, 11, 8, 8, 9])
    keywords.freeze_panes = "A2"
    keywords.auto_filter.ref = keywords.dimensions

    clusters_sheet.append(["Cluster", "Label", "Intent", "Phrases", "Priority",
                           "Head phrase", "Go to keywords"])
    for cluster in clusters:
        clusters_sheet.append([cluster.index, cluster.label, cluster.intent,
                               cluster.size, round(cluster.priority, 1),
                               cluster.head.text, "open ->"])
        row = clusters_sheet.max_row
        # An in-workbook jump is a `location`, NOT a target. Assigning a plain
        # "#Sheet!A1" string makes openpyxl write an EXTERNAL relationship whose
        # target merely starts with "#", which is malformed: LibreOffice ignores
        # it outright and Excel only resolves it by leniency. Building the
        # Hyperlink with location= and no target emits the standard form -
        # <hyperlink ref="A2" location="'Keywords'!A2"/> with no relationship -
        # which every spreadsheet reads. The sheet name is quoted so a name with
        # a space would not break it either.
        location = f"'Keywords'!A{first_row[cluster.index]}"
        for column in ("A", "G"):
            cell = clusters_sheet[f"{column}{row}"]
            cell.hyperlink = Hyperlink(ref=f"{column}{row}", location=location)
            cell.font = link_font
    style_header(clusters_sheet, [9, 32, 15, 10, 10, 52, 16])
    clusters_sheet.freeze_panes = "A2"
    clusters_sheet.auto_filter.ref = f"A1:F{clusters_sheet.max_row}"

    complete = ("closed - nothing new was left to find" if meta.get("exhausted")
                else "stopped at the time limit" if meta.get("stopped_by_time_limit")
                else "stopped by a rate limit" if meta.get("stopped_by_rate_limit")
                else "stopped at the level or query limit")
    geography = (f"mixed across a rotating proxy pool"
                 if meta.get("egress_country") == "mixed"
                 else f"{meta.get('egress_country', '')} (the exit IP decides it)")

    run.append(["Keyword universe", meta.get("seed", "")])
    run["A1"].font = title_font
    run["B1"].font = title_font
    run.append([])
    run.append(["Summary", ""])
    summary_header = run.max_row
    for label, value in (
        ("Keywords found", f"{meta.get('keywords', meta.get('phrases', 0)):,}"),
        ("Phrases Google returned", f"{meta.get('phrases', 0):,}"),
        ("Collapsed as re-worded duplicates", f"{meta.get('variants_collapsed', 0):,}"),
        ("Topic clusters", f"{meta.get('clusters', 0):,}"),
        ("Harvested (UTC)", meta.get("harvested_at", "")),
        ("Source", "Google autocomplete only"),
        ("Market / geography", geography),
        ("Language", meta.get("language", "")),
        ("How it ended", complete),
        ("Rounds completed", meta.get("levels_run", "")),
        ("Phrases not yet expanded", f"{meta.get('unexpanded_phrases', 0):,}"),
        ("Queries asked", f"{meta.get('queries_asked', 0):,}"),
        ("Of those, served from cache", f"{meta.get('cache_hits', 0):,}"),
        ("Time taken", f"{meta.get('elapsed_seconds', 0):,.0f}s"),
    ):
        run.append([label, str(value)])

    run.append([])
    run.append(["What the columns mean", ""])
    meanings_header = run.max_row
    for column, meaning in COLUMN_MEANINGS:
        run.append([column, meaning])
        run[f"A{run.max_row}"].font = Font(bold=True)
        run[f"B{run.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")

    run.column_dimensions["A"].width = 28
    run.column_dimensions["B"].width = 104
    for header_row in (summary_header, meanings_header):
        for cell in run[header_row]:
            cell.font = header_font
            cell.fill = header_fill

    contamination = sorted(universe.off_seed.items(), key=lambda kv: -kv[1])
    if contamination:
        off = book.create_sheet("Off-seed")
        off.append(["Times returned", "Phrase Google returned that lacks the seed"])
        for phrase, count in contamination[:200]:
            off.append([count, phrase])
        style_header(off, [16, 60])
        off.freeze_panes = "A2"

    book.active = 0
    book.save(path)
    return True


def write_all(outdir: str, universe: Universe, clusters: list[Cluster],
              meta: dict) -> dict[str, str]:
    os.makedirs(outdir, exist_ok=True)
    stem = "".join(c if c.isalnum() else "-" for c in universe.seed.lower()).strip("-")
    paths = {
        "json": os.path.join(outdir, f"{stem}.json"),
        "csv": os.path.join(outdir, f"{stem}.csv"),
        "md": os.path.join(outdir, f"{stem}.md"),
    }
    write_json(paths["json"], universe, clusters, meta)
    write_csv(paths["csv"], clusters)
    write_markdown(paths["md"], universe, clusters, meta)

    xlsx_path = os.path.join(outdir, f"{stem}.xlsx")
    if write_xlsx(xlsx_path, universe, clusters, meta):
        paths["xlsx"] = xlsx_path
    return paths
